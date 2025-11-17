import torch
import numpy as np

def bandpower(signal, fs, f_low, f_high):
    """
    Compute bandpower via FFT magnitude integration.
    signal: array [B, T] or [T]
    fs: sampling rate
    f_low, f_high: frequency range
    """
    if isinstance(signal, torch.Tensor):
        x = signal.detach().cpu().numpy()
    else:
        x = np.asarray(signal)
    if x.ndim == 1:
        x = x[None, :]
    B, T = x.shape

    freqs = np.fft.rfftfreq(T, d=1.0/fs)
    X = np.fft.rfft(x, axis=1)
    psd = (np.abs(X) ** 2) / T
    idx = (freqs >= f_low) & (freqs <= f_high)
    bp = psd[:, idx].sum(axis=1)
    return bp  # shape [B]

def delta_percent(after, before):
    """Percentage change (after vs before)."""
    before = np.asarray(before)
    after = np.asarray(after)
    return 100.0 * (after - before) / (before + 1e-20)

def compute_metrics(raw, denoised, fs):
    """
    Compute EEG denoising metrics.
    raw, denoised: arrays/tensors [B, T] or [T]
    fs: sampling rate
    Returns dict with metrics arrays per item.
    """
    if isinstance(raw, torch.Tensor):
        raw = raw.detach().cpu().numpy()
    if isinstance(denoised, torch.Tensor):
        denoised = denoised.detach().cpu().numpy()
    if raw.ndim == 1:
        raw = raw[None, :]
    if denoised.ndim == 1:
        denoised = denoised[None, :]
    B, T = raw.shape

    # Bandpowers
    eeg_raw = bandpower(raw, fs, 0.5, 40.0)
    hf_raw  = bandpower(raw, fs, 80.0, min(500.0, fs/2 - 1e-6))
    eeg_den = bandpower(denoised, fs, 0.5, 40.0)
    hf_den  = bandpower(denoised, fs, 80.0, min(500.0, fs/2 - 1e-6))

    # SNR in dB
    snr_raw = np.log10((eeg_raw + 1e-20) / (hf_raw + 1e-20)) * 10.0
    snr_den = np.log10((eeg_den + 1e-20) / (hf_den + 1e-20)) * 10.0
    delta_snr_db = snr_den - snr_raw

    # Total EEG (0.5–500 Hz)
    eeg_total_raw = bandpower(raw, fs, 0.5, min(500.0, fs/2 - 1e-6))
    eeg_total_den = bandpower(denoised, fs, 0.5, min(500.0, fs/2 - 1e-6))
    eeg_total_delta = delta_percent(eeg_total_den, eeg_total_raw)

    # Low EEG (0.5–40 Hz)
    eeg_low_delta   = delta_percent(eeg_den, eeg_raw)

    # HF (80–500 Hz) suppression
    hf_suppress     = delta_percent(hf_den, hf_raw)  # negative = suppression

    return {
        "SNR_raw_dB": snr_raw,
        "SNR_dip_dB": snr_den,
        "Delta_SNR_dB": delta_snr_db,
        "EEG_total_delta_pct": eeg_total_delta,
        "EEG_0p5_40_delta_pct": eeg_low_delta,
        "HF_suppress_pct": hf_suppress,
    }

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def export_presentation_table(csv_path: str):
    """
    Convert numeric DIP EEG results into a professional presentation table.
    Splits HF into HF1/HF2, highlights low-frequency EEG change and HF suppression,
    and renames 'model' -> 'prior' for research presentation.
    """
    df = pd.read_csv(csv_path)

    # Compute rounded summary metrics
    df["EEG_low_change_%"] = df["EEG_0p5_40_delta_pct"].round(2)
    df["HF1_suppress_%"]   = df["HF1_suppress_pct"].round(2)
    df["HF2_suppress_%"]   = df["HF2_suppress_pct"].round(2)
    df["ΔSNR_dB"]          = df["Delta_SNR_dB"].round(2)

    # Rename model → prior
    df.rename(columns={"model": "prior"}, inplace=True)

    # Choose key columns for presentation
    df = df[[
        "label","channel","is_artifact",
        "EEG_low_change_%","HF1_suppress_%","HF2_suppress_%","ΔSNR_dB",
        "EEG_total_0p5_500_raw","EEG_total_0p5_500_dip",
        "selection","agg","prior"
    ]]

    # Rename columns for publication clarity
    df.rename(columns={
        "label": "Sleep Stage",
        "channel": "Channel",
        "is_artifact": "Artifact Window",
        "EEG_low_change_%": "Δ EEG (0.5–40 Hz) [%]",
        "HF1_suppress_%": "HF1 Suppression (80–250 Hz) [%]",
        "HF2_suppress_%": "HF2 Suppression (250–500 Hz) [%]",
        "ΔSNR_dB": "Δ SNR [dB]",
        "EEG_total_0p5_500_raw": "EEG Power Raw [V²]",
        "EEG_total_0p5_500_dip": "EEG Power Denoised [V²]",
        "selection": "Criterion",
        "agg": "Aggregation",
        "prior": "Prior"
    }, inplace=True)

    # Export to Excel
    xlsx_path = csv_path.replace(".csv", "_presentable.xlsx")
    df.to_excel(xlsx_path, index=False)

    # Apply simple Excel header styling
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(col_name) + 2)
    wb.save(xlsx_path)
    wb.close()

    print(f"[OK] Presentation-ready table saved: {xlsx_path}")